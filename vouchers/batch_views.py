"""
Batch Signature System Views
Allows Finance Manager to group approved vouchers and send to MD for bulk signature
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from decimal import Decimal
from io import BytesIO

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import (
    PaymentVoucher, PaymentForm,
    SignatureBatch, BatchVoucherItem, BatchFormItem
)


# ============================================================================
# FINANCE MANAGER: BATCH SELECTION
# ============================================================================

@login_required
def batch_select_documents(request):
    """
    Account Payable or Finance Manager selects PENDING_L5 vouchers/forms for batch signature
    Accessible by Account Payable (role_level 1) and Finance Manager (role_level 3)
    Documents at PENDING_L5 status means GM has approved and waiting for MD
    """
    # Check permission - Account Payable or Finance Manager
    if request.user.role_level not in [1, 3]:
        messages.error(request, 'Only Account Payable or Finance Managers can create signature batches')
        return redirect('dashboard:home')

    # Get IDs of documents already in PENDING batches to exclude them
    pending_voucher_ids = BatchVoucherItem.objects.filter(
        batch__status='PENDING'
    ).values_list('voucher_id', flat=True)

    pending_form_ids = BatchFormItem.objects.filter(
        batch__status='PENDING'
    ).values_list('payment_form_id', flat=True)

    # Get ALL PENDING_L5 documents (after GM approval, waiting for MD)
    # Exclude documents already in a pending batch
    vouchers = PaymentVoucher.objects.filter(
        status='PENDING_L5'
    ).exclude(
        id__in=pending_voucher_ids
    ).select_related('created_by').prefetch_related('line_items').order_by('-created_at')

    forms = PaymentForm.objects.filter(
        status='PENDING_L5'
    ).exclude(
        id__in=pending_form_ids
    ).select_related('created_by').prefetch_related('line_items').order_by('-created_at')

    context = {
        'vouchers': vouchers,
        'forms': forms,
    }

    return render(request, 'vouchers/batch/select_documents.html', context)


@login_required
def batch_create(request):
    """
    Create a new signature batch from selected documents
    """
    # Check permission - Account Payable or Finance Manager
    if request.user.role_level not in [1, 3]:
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    if request.method == 'POST':
        # Get selected voucher and form IDs
        voucher_ids = request.POST.getlist('voucher_ids[]')
        form_ids = request.POST.getlist('form_ids[]')
        fm_notes = request.POST.get('notes', '')

        # Validate at least one document selected
        if not voucher_ids and not form_ids:
            return JsonResponse({'error': 'Please select at least one document'}, status=400)

        # Create batch
        batch = SignatureBatch.objects.create(
            created_by=request.user,
            fm_notes=fm_notes
        )

        # Add vouchers to batch (must be PENDING_L5 - after GM approval)
        for voucher_id in voucher_ids:
            try:
                voucher = PaymentVoucher.objects.get(
                    id=voucher_id,
                    status='PENDING_L5'
                )
                BatchVoucherItem.objects.create(batch=batch, voucher=voucher)
            except PaymentVoucher.DoesNotExist:
                pass

        # Add forms to batch (must be PENDING_L5 - after GM approval)
        for form_id in form_ids:
            try:
                form = PaymentForm.objects.get(
                    id=form_id,
                    status='PENDING_L5'
                )
                BatchFormItem.objects.create(batch=batch, payment_form=form)
            except PaymentForm.DoesNotExist:
                pass

        # Add success message
        messages.success(request, f'Batch {batch.batch_number} created successfully with {batch.get_document_count()} document(s)!')

        return JsonResponse({
            'success': True,
            'batch_number': batch.batch_number,
            'document_count': batch.get_document_count(),
            'redirect_url': f'/vouchers/batch/{batch.id}/detail/'
        })

    return JsonResponse({'error': 'Invalid request method'}, status=400)


# ============================================================================
# MD DASHBOARD
# ============================================================================

@login_required
def fm_batch_list(request):
    """
    View to track batches created by current user
    Accessible by Account Payable (role_level 1) and Finance Manager (role_level 3)
    """
    # Check permission - Account Payable or Finance Manager
    if request.user.role_level not in [1, 3]:
        messages.error(request, 'Only Account Payable or Finance Managers can access this page')
        return redirect('dashboard:home')

    # Get all batches created by this user
    batches = SignatureBatch.objects.filter(
        created_by=request.user
    ).select_related('created_by', 'signed_by').prefetch_related(
        'voucher_items__voucher',
        'form_items__payment_form'
    ).order_by('-created_at')

    # Separate by status
    pending_batches = batches.filter(status='PENDING')
    signed_batches = batches.filter(status='SIGNED')
    rejected_batches = batches.filter(status='REJECTED')

    context = {
        'pending_batches': pending_batches,
        'signed_batches': signed_batches,
        'rejected_batches': rejected_batches,
        'pending_count': pending_batches.count(),
        'signed_count': signed_batches.count(),
        'rejected_count': rejected_batches.count(),
    }

    return render(request, 'vouchers/batch/fm_batch_list.html', context)


@login_required
def md_dashboard(request):
    """
    Managing Director dashboard to view and sign batches
    Only accessible by MD (role_level 5)
    """
    # Check permission
    if request.user.role_level != 5:
        messages.error(request, 'Only Managing Director can access this dashboard')
        return redirect('dashboard:home')

    # Get pending batches
    pending_batches = SignatureBatch.objects.filter(
        status='PENDING'
    ).select_related('created_by').prefetch_related(
        'voucher_items__voucher',
        'form_items__payment_form'
    ).order_by('-created_at')

    # Get signed batches (last 10)
    signed_batches = SignatureBatch.objects.filter(
        status='SIGNED'
    ).select_related('created_by', 'signed_by').order_by('-signed_at')[:10]

    context = {
        'pending_batches': pending_batches,
        'signed_batches': signed_batches,
        'pending_count': pending_batches.count(),
    }

    return render(request, 'vouchers/batch/md_dashboard.html', context)


@login_required
def all_batches_list(request):
    """
    View all signature batches in the system
    All authenticated users can view for transparency
    """
    # Get all batches
    batches = SignatureBatch.objects.select_related(
        'created_by', 'signed_by'
    ).prefetch_related(
        'voucher_items__voucher',
        'form_items__payment_form'
    ).order_by('-created_at')

    # Separate by status
    pending_batches = batches.filter(status='PENDING')
    signed_batches = batches.filter(status='SIGNED')
    rejected_batches = batches.filter(status='REJECTED')

    context = {
        'pending_batches': pending_batches,
        'signed_batches': signed_batches,
        'rejected_batches': rejected_batches,
        'pending_count': pending_batches.count(),
        'signed_count': signed_batches.count(),
        'rejected_count': rejected_batches.count(),
        'total_count': batches.count(),
    }

    return render(request, 'vouchers/batch/all_batches_list.html', context)


@login_required
def batch_detail(request, batch_id):
    """
    View details of a signature batch
    All authenticated users can view batches for transparency
    """
    batch = get_object_or_404(
        SignatureBatch.objects.prefetch_related(
            'voucher_items__voucher__line_items',
            'form_items__payment_form__line_items'
        ),
        id=batch_id
    )

    # All authenticated users can view batch details (read-only for L2, L4)
    # No permission check - transparency for all users

    context = {
        'batch': batch,
        'total_amount': batch.get_total_amount(),
        'total_amount_display': batch.get_total_amount_display(),
        'document_count': batch.get_document_count(),
    }

    return render(request, 'vouchers/batch/batch_detail.html', context)


@login_required
def batch_sign(request, batch_id):
    """
    MD signs all documents in a batch
    """
    # Check permission
    if request.user.role_level != 5:
        return JsonResponse({'error': 'Only MD can sign batches'}, status=403)

    batch = get_object_or_404(SignatureBatch, id=batch_id, status='PENDING')

    if request.method == 'POST':
        md_comments = request.POST.get('comments', '')

        # Get IP address
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip_address = x_forwarded_for.split(',')[0]
        else:
            ip_address = request.META.get('REMOTE_ADDR')

        # Approve all documents in batch using state machine
        from workflow.state_machine import VoucherStateMachine, FormStateMachine

        comment_text = f'✓ Approved via Batch {batch.batch_number}'
        if md_comments:
            comment_text += f' - {md_comments}'

        success_count = 0
        error_count = 0
        errors = []

        # Approve all vouchers in batch
        for item in batch.voucher_items.all():
            try:
                voucher = item.voucher
                # Validate the voucher is at PENDING_L5
                if voucher.status == 'PENDING_L5':
                    VoucherStateMachine.transition(voucher, 'approve', request.user, comment_text, via_batch=True)
                    success_count += 1
                else:
                    errors.append(f"PV {voucher.pv_number}: Invalid status ({voucher.get_status_display()})")
                    error_count += 1
            except Exception as e:
                errors.append(f"PV {voucher.pv_number}: {str(e)}")
                error_count += 1

        # Approve all forms in batch
        for item in batch.form_items.all():
            try:
                form = item.payment_form
                # Validate the form is at PENDING_L5
                if form.status == 'PENDING_L5':
                    FormStateMachine.transition(form, 'approve', request.user, comment_text, via_batch=True)
                    success_count += 1
                else:
                    errors.append(f"PF {form.pf_number}: Invalid status ({form.get_status_display()})")
                    error_count += 1
            except Exception as e:
                errors.append(f"PF {form.pf_number}: {str(e)}")
                error_count += 1

        # Update batch status
        batch.status = 'SIGNED'
        batch.signed_by = request.user
        batch.signed_at = timezone.now()
        batch.signature_ip = ip_address
        batch.md_comments = md_comments
        batch.save()

        # Add success/error messages
        if error_count > 0:
            messages.warning(request, f'Batch {batch.batch_number} processed with {error_count} error(s). {success_count} document(s) approved successfully.')
            for error in errors[:5]:  # Show first 5 errors
                messages.error(request, error)
        else:
            messages.success(request, f'Batch {batch.batch_number} signed successfully! All {success_count} document(s) approved.')

        return JsonResponse({
            'success': True,
            'message': f'Successfully approved {success_count} document(s) in batch {batch.batch_number}',
            'errors': errors if error_count > 0 else [],
            'error_count': error_count,
            'redirect_url': '/vouchers/md-dashboard/'
        })

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def batch_reject(request, batch_id):
    """
    MD rejects a batch
    """
    # Check permission
    if request.user.role_level != 5:
        return JsonResponse({'error': 'Only MD can reject batches'}, status=403)

    batch = get_object_or_404(SignatureBatch, id=batch_id, status='PENDING')

    if request.method == 'POST':
        md_comments = request.POST.get('comments', '')

        if not md_comments:
            return JsonResponse({'error': 'Please provide rejection reason'}, status=400)

        # Update batch
        batch.status = 'REJECTED'
        batch.signed_by = request.user
        batch.signed_at = timezone.now()
        batch.md_comments = md_comments
        batch.save()

        # Add warning message
        messages.warning(request, f'Batch {batch.batch_number} rejected')

        return JsonResponse({
            'success': True,
            'message': f'Batch {batch.batch_number} rejected. Finance Manager will be notified.',
            'redirect_url': '/vouchers/md-dashboard/'
        })

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def batch_remove_document(request, batch_id):
    """
    MD removes a single document from a pending batch
    """
    # Check permission
    if request.user.role_level != 5:
        return JsonResponse({'error': 'Only MD can remove documents from batches'}, status=403)

    batch = get_object_or_404(SignatureBatch, id=batch_id, status='PENDING')

    if request.method == 'POST':
        doc_type = request.POST.get('doc_type')  # 'voucher' or 'form'
        doc_id = request.POST.get('doc_id')
        reason = request.POST.get('reason', '')

        if not reason:
            return JsonResponse({'error': 'Please provide a reason for removal'}, status=400)

        try:
            if doc_type == 'voucher':
                item = BatchVoucherItem.objects.get(batch=batch, voucher_id=doc_id)
                doc_number = item.voucher.pv_number
                item.delete()
            elif doc_type == 'form':
                item = BatchFormItem.objects.get(batch=batch, payment_form_id=doc_id)
                doc_number = item.payment_form.pf_number
                item.delete()
            else:
                return JsonResponse({'error': 'Invalid document type'}, status=400)

            # Check if batch is now empty
            if batch.get_document_count() == 0:
                batch.status = 'REJECTED'
                batch.signed_by = request.user
                batch.signed_at = timezone.now()
                batch.md_comments = f'All documents removed. Last removal reason: {reason}'
                batch.save()

                return JsonResponse({
                    'success': True,
                    'batch_empty': True,
                    'message': f'Document {doc_number} removed. Batch is now empty and marked as rejected.',
                    'redirect_url': '/vouchers/md-dashboard/'
                })

            # Recalculate total
            new_total = batch.get_total_amount_display()
            new_count = batch.get_document_count()

            return JsonResponse({
                'success': True,
                'batch_empty': False,
                'message': f'Document {doc_number} removed from batch',
                'new_total': new_total,
                'new_count': new_count
            })

        except (BatchVoucherItem.DoesNotExist, BatchFormItem.DoesNotExist):
            return JsonResponse({'error': 'Document not found in batch'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def batch_export_excel(request, batch_id):
    """
    Export a signature batch to Excel format
    Shows all documents in the batch with details
    All authenticated users can export for transparency and record keeping
    """
    batch = get_object_or_404(
        SignatureBatch.objects.prefetch_related(
            'voucher_items__voucher__line_items__department',
            'form_items__payment_form__line_items__department'
        ),
        id=batch_id
    )

    # All authenticated users can export batches
    # No permission check - transparency for all users

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Batch {batch.batch_number}"

    # Column widths
    column_widths = {
        'A': 3.71,   # No
        'B': 15.14,  # Doc No
        'C': 25.0,   # Supplier
        'D': 50.0,   # Description
        'E': 18.0,   # Amount
        'F': 28.0,   # Transfer Account
        'G': 25.0,   # Account Name
        'H': 20.0,   # Account Number
        'I': 20.0    # Remark
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    company_font = Font(name='Calibri', size=14, bold=True, color='000000')
    title_font = Font(name='Calibri', size=13, bold=True, color='000000')
    data_font = Font(name='Calibri', size=11, color='000000')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center')

    # Row 1: Company name
    ws.merge_cells('A1:I1')
    company_cell = ws['A1']
    company_cell.value = "Phat Phnom Penh  Co.,LTD  (Garden City Water Park)"
    company_cell.font = company_font
    company_cell.alignment = center_alignment

    # Row 2: Title
    ws.merge_cells('A2:I2')
    title_cell = ws['A2']
    title_cell.value = f"Signature Batch - {batch.batch_number}"
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # Row 3: Batch info
    batch_info = f"Created by: {batch.created_by.get_full_name() or batch.created_by.username} | Date: {batch.created_at.strftime('%Y-%m-%d')} | Status: {batch.get_status_display()}"
    if batch.status == 'SIGNED' and batch.signed_by:
        batch_info += f" | Signed by: {batch.signed_by.get_full_name() or batch.signed_by.username} on {batch.signed_at.strftime('%Y-%m-%d')}"

    ws.merge_cells('A3:I3')
    info_cell = ws['A3']
    info_cell.value = batch_info
    info_cell.font = Font(name='Calibri', size=10, color='666666')
    info_cell.alignment = left_alignment

    # Row 4: FM Notes (if any)
    if batch.fm_notes:
        ws.merge_cells('A4:I4')
        notes_cell = ws['A4']
        notes_cell.value = f"FM Notes: {batch.fm_notes}"
        notes_cell.font = Font(name='Calibri', size=10, italic=True)
        notes_cell.alignment = left_alignment
        header_row_start = 6
    else:
        header_row_start = 5

    # Header rows
    merge_ranges = [
        f'A{header_row_start}:A{header_row_start+1}',
        f'B{header_row_start}:B{header_row_start+1}',
        f'C{header_row_start}:C{header_row_start+1}',
        f'D{header_row_start}:D{header_row_start+1}',
        f'E{header_row_start}:E{header_row_start+1}',
        f'F{header_row_start}:F{header_row_start+1}',
        f'G{header_row_start}:H{header_row_start}',
        f'I{header_row_start}:I{header_row_start+1}'
    ]

    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)

    # Header row 1
    headers_row1 = [
        (f'A{header_row_start}', 'No'),
        (f'B{header_row_start}', 'Doc No.'),
        (f'C{header_row_start}', 'Supplier'),
        (f'D{header_row_start}', 'Description'),
        (f'E{header_row_start}', 'Amount'),
        (f'F{header_row_start}', 'Transfer Account'),
        (f'G{header_row_start}', 'Receiver Bank Account'),
        (f'I{header_row_start}', 'Remark')
    ]

    for cell_ref, value in headers_row1:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = border

    # Header row 2 (sub-headers)
    headers_row2 = [
        (f'G{header_row_start+1}', 'Account Name'),
        (f'H{header_row_start+1}', 'Account Number')
    ]

    for cell_ref, value in headers_row2:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = border

    # Apply borders to all header cells
    for row in [header_row_start, header_row_start+1]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = ws[f'{col}{row}']
            if not cell.border.left.style:
                cell.border = border
                cell.fill = header_fill

    # Data rows
    row = header_row_start + 2
    total_amount = Decimal('0')
    idx = 0

    # Collect all documents (vouchers and forms)
    all_documents = []

    for item in batch.voucher_items.all():
        all_documents.append(('PV', item.voucher))

    for item in batch.form_items.all():
        all_documents.append(('PF', item.payment_form))

    # Sort by created date
    all_documents.sort(key=lambda x: x[1].created_at)

    for doc_type, doc in all_documents:
        idx += 1

        doc_number = doc.pv_number if doc_type == 'PV' else doc.pf_number

        # Get transfer account
        if doc.company_bank_account:
            transfer_account = doc.company_bank_account.bank
        else:
            transfer_account = "No Transfer Account"

        # Calculate totals
        totals = doc.calculate_grand_total()

        # Combine descriptions
        descriptions = [item.description for item in doc.line_items.all()]
        combined_description = '\n'.join(descriptions) if descriptions else ''

        # Calculate USD amount
        amount = totals.get('USD', Decimal('0'))
        total_amount += amount

        # Payee bank info
        account_name = doc.bank_name or ''
        account_number = doc.bank_account_number or ''

        # Column A: No
        cell_a = ws.cell(row=row, column=1)
        cell_a.value = idx
        cell_a.font = data_font
        cell_a.alignment = center_alignment
        cell_a.border = border

        # Column B: Doc Number
        cell_b = ws.cell(row=row, column=2)
        cell_b.value = doc_number or 'DRAFT'
        cell_b.font = data_font
        cell_b.alignment = left_alignment
        cell_b.border = border

        # Column C: Supplier
        cell_c = ws.cell(row=row, column=3)
        cell_c.value = doc.payee_name
        cell_c.font = data_font
        cell_c.alignment = left_alignment
        cell_c.border = border

        # Column D: Description
        cell_d = ws.cell(row=row, column=4)
        cell_d.value = combined_description
        cell_d.font = data_font
        cell_d.alignment = left_alignment
        cell_d.border = border

        # Column E: Amount
        cell_e = ws.cell(row=row, column=5)
        cell_e.value = float(amount)
        cell_e.font = data_font
        cell_e.alignment = right_alignment
        cell_e.border = border
        cell_e.number_format = '#,##0.00'

        # Column F: Transfer Account
        cell_f = ws.cell(row=row, column=6)
        cell_f.value = transfer_account
        cell_f.font = data_font
        cell_f.alignment = left_alignment
        cell_f.border = border

        # Column G: Account Name
        cell_g = ws.cell(row=row, column=7)
        cell_g.value = account_name
        cell_g.font = data_font
        cell_g.alignment = left_alignment
        cell_g.border = border

        # Column H: Account Number
        cell_h = ws.cell(row=row, column=8)
        cell_h.value = account_number
        cell_h.font = data_font
        cell_h.alignment = left_alignment
        cell_h.border = border

        # Column I: Remark
        cell_i = ws.cell(row=row, column=9)
        cell_i.value = ''
        cell_i.font = data_font
        cell_i.alignment = left_alignment
        cell_i.border = border

        row += 1

    # Grand total row
    total_label_cell = ws.cell(row=row, column=4)
    total_label_cell.value = "GRAND TOTAL:"
    total_label_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    total_label_cell.alignment = right_alignment
    total_label_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_label_cell.border = border

    total_amount_cell = ws.cell(row=row, column=5)
    total_amount_cell.value = float(total_amount)
    total_amount_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    total_amount_cell.alignment = right_alignment
    total_amount_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    total_amount_cell.border = border
    total_amount_cell.number_format = '#,##0.00'

    for col in [1, 2, 3, 6, 7, 8, 9]:
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    row += 3

    # Signature section
    ws.merge_cells(f'B{row}:C{row}')
    prepared_cell = ws.cell(row=row, column=2)
    prepared_cell.value = "Prepared by"
    prepared_cell.font = header_font
    prepared_cell.alignment = center_alignment

    ws.merge_cells(f'E{row}:F{row}')
    checked_cell = ws.cell(row=row, column=5)
    checked_cell.value = "Checked by"
    checked_cell.font = header_font
    checked_cell.alignment = center_alignment

    ws.merge_cells(f'H{row}:I{row}')
    approved_cell = ws.cell(row=row, column=8)
    approved_cell.value = "Approved by"
    approved_cell.font = header_font
    approved_cell.alignment = center_alignment

    # Page setup
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToWidth = 1
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Save to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Batch_{batch.batch_number}_{timezone.now().strftime("%Y%m%d")}.xlsx'

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response