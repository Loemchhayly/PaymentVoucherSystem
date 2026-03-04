"""
Advanced Reporting System for Payment Vouchers and Forms
Supports Excel and PDF exports with comprehensive filters
"""
from django.http import HttpResponse
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime, timedelta
from io import BytesIO
from decimal import Decimal

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from .models import PaymentVoucher, PaymentForm, Department
from accounts.models import User


class ReportGenerator:
    """Main report generator class with advanced filtering"""

    def __init__(self, filters=None):
        """Initialize with optional filters"""
        self.filters = filters or {}
        self.vouchers = None
        self.forms = None

    def apply_filters(self):
        """Apply filters to querysets"""
        # Start with all vouchers and forms
        vouchers_qs = PaymentVoucher.objects.all()
        forms_qs = PaymentForm.objects.all()

        # Date range filter
        date_from = self.filters.get('date_from')
        date_to = self.filters.get('date_to')

        if date_from:
            vouchers_qs = vouchers_qs.filter(payment_date__gte=date_from)
            forms_qs = forms_qs.filter(payment_date__gte=date_from)

        if date_to:
            vouchers_qs = vouchers_qs.filter(payment_date__lte=date_to)
            forms_qs = forms_qs.filter(payment_date__lte=date_to)

        # Status filter - default to show documents after L2 (Account Supervisor) approval
        # This includes: PENDING_L3, PENDING_L4, PENDING_L5, and APPROVED
        status = self.filters.get('status')
        if status and status != 'ALL':
            vouchers_qs = vouchers_qs.filter(status=status)
            forms_qs = forms_qs.filter(status=status)
        else:
            # If no status filter specified, show documents after L2 approval
            # (PENDING_L3 onwards means L2 has already approved)
            allowed_statuses = ['PENDING_L3', 'PENDING_L4', 'PENDING_L5', 'APPROVED']
            vouchers_qs = vouchers_qs.filter(status__in=allowed_statuses)
            forms_qs = forms_qs.filter(status__in=allowed_statuses)

        # Creator filter
        creator_id = self.filters.get('creator')
        if creator_id:
            vouchers_qs = vouchers_qs.filter(created_by_id=creator_id)
            forms_qs = forms_qs.filter(created_by_id=creator_id)

        # Department filter
        department_name = self.filters.get('department')
        if department_name:
            vouchers_qs = vouchers_qs.filter(line_items__department__name=department_name).distinct()
            forms_qs = forms_qs.filter(line_items__department__name=department_name).distinct()

        # Payee filter
        payee_name = self.filters.get('payee_name')
        if payee_name:
            vouchers_qs = vouchers_qs.filter(payee_name__icontains=payee_name)
            forms_qs = forms_qs.filter(payee_name__icontains=payee_name)

        # Document type filter
        doc_type = self.filters.get('doc_type', 'ALL')
        if doc_type == 'PV':
            forms_qs = PaymentForm.objects.none()
        elif doc_type == 'PF':
            vouchers_qs = PaymentVoucher.objects.none()

        self.vouchers = vouchers_qs.select_related('created_by').prefetch_related('line_items')
        self.forms = forms_qs.select_related('created_by').prefetch_related('line_items')

        return self

    def get_summary_stats(self):
        """Calculate summary statistics"""
        if self.vouchers is None or self.forms is None:
            self.apply_filters()

        stats = {
            'total_vouchers': self.vouchers.count(),
            'total_forms': self.forms.count(),
            'total_documents': self.vouchers.count() + self.forms.count(),
            'by_status': {},
            'by_currency': {'USD': Decimal('0'), 'KHR': Decimal('0'), 'THB': Decimal('0')},
        }

        # Count by status
        for status_code, status_label in PaymentVoucher.STATUS_CHOICES:
            count = self.vouchers.filter(status=status_code).count() + self.forms.filter(status=status_code).count()
            if count > 0:
                stats['by_status'][status_label] = count

        # Calculate totals by currency
        for voucher in self.vouchers:
            totals = voucher.calculate_grand_total()
            for currency, amount in totals.items():
                stats['by_currency'][currency] += amount

        for form in self.forms:
            totals = form.calculate_grand_total()
            for currency, amount in totals.items():
                stats['by_currency'][currency] += amount

        return stats

    def export_to_excel(self):
        """Export filtered data to Excel grouped by bank account sections"""
        if self.vouchers is None or self.forms is None:
            self.apply_filters()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Summary by Bank"

        # ===========================================================================
        # SET COLUMN WIDTHS
        # ===========================================================================
        column_widths = {
            'A': 5.0,    # No
            'B': 15.0,   # PV/PF No.
            'C': 25.0,   # Supplier
            'D': 45.0,   # Description
            'E': 15.0,   # Amount
            'F': 25.0,   # Receiver Account Name
            'G': 20.0,   # Receiver Account Number
            'H': 20.0    # Remark
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # ===========================================================================
        # DEFINE STYLES
        # ===========================================================================
        header_font = Font(name='Calibri', size=11, bold=True, color='000000')
        company_font = Font(name='Calibri', size=14, bold=True, color='000000')
        title_font = Font(name='Calibri', size=13, bold=True, color='000000')
        data_font = Font(name='Calibri', size=11, color='000000')

        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center')

        # ===========================================================================
        # ROW 1: COMPANY NAME
        # ===========================================================================
        ws.merge_cells('A1:H1')
        company_cell = ws['A1']
        company_cell.value = "Phat Phnom Penh  Co.,LTD  (Garden City Water Park)"
        company_cell.font = company_font
        company_cell.alignment = center_alignment

        # ===========================================================================
        # ROW 2: TITLE & DATE
        # ===========================================================================
        ws.merge_cells('A2:F2')
        title_cell = ws['A2']
        title_cell.value = "Payment Summary by Bank Account"
        title_cell.font = title_font
        title_cell.alignment = center_alignment

        # Date on the right (G2:H2)
        ws.merge_cells('G2:H2')
        date_cell = ws['G2']
        date_cell.value = f"Date: {timezone.now().strftime('%d-%b-%Y')}"
        date_cell.font = Font(name='Calibri', size=11, bold=True)
        date_cell.alignment = right_alignment

        # ===========================================================================
        # GROUP DOCUMENTS BY BANK AND TYPE
        # ===========================================================================
        # Define bank sections
        bank_sections = [
            {'code': 'AC', 'name': 'ACLEDA Bank', 'type': 'PV', 'title': 'AC PV'},
            {'code': 'ABA', 'name': 'ABA Bank', 'type': 'PV', 'title': 'ABA PV'},
            {'code': 'AC', 'name': 'ACLEDA Bank', 'type': 'PF', 'title': 'AC PF'},
            {'code': 'ABA', 'name': 'ABA Bank', 'type': 'PF', 'title': 'ABA PF'},
        ]

        # Group documents by bank and type
        def categorize_document(doc):
            """Return bank code and document type"""
            is_voucher = hasattr(doc, 'pv_number')
            doc_type = 'PV' if is_voucher else 'PF'

            if doc.company_bank_account:
                bank_name = doc.company_bank_account.bank
                if 'ACLEDA' in bank_name.upper() or 'AC' in bank_name.upper():
                    return 'AC', doc_type
                elif 'ABA' in bank_name.upper():
                    return 'ABA', doc_type
            return None, doc_type

        # Organize documents into sections
        sections_data = {
            ('AC', 'PV'): [],
            ('ABA', 'PV'): [],
            ('AC', 'PF'): [],
            ('ABA', 'PF'): [],
        }

        for voucher in self.vouchers:
            bank_code, doc_type = categorize_document(voucher)
            if bank_code and (bank_code, doc_type) in sections_data:
                sections_data[(bank_code, doc_type)].append(voucher)

        for form in self.forms:
            bank_code, doc_type = categorize_document(form)
            if bank_code and (bank_code, doc_type) in sections_data:
                sections_data[(bank_code, doc_type)].append(form)

        # ===========================================================================
        # START BUILDING SECTIONS
        # ===========================================================================
        row = 4  # Start after title and date
        grand_totals = {'USD': Decimal('0'), 'KHR': Decimal('0'), 'THB': Decimal('0')}

        # ===========================================================================
        # GENERATE SECTIONS
        # ===========================================================================
        # Professional color scheme
        section_header_fill = PatternFill(start_color='FDB462', end_color='FDB462', fill_type='solid')  # Professional orange
        table_header_fill = PatternFill(start_color='B3CDE3', end_color='B3CDE3', fill_type='solid')   # Professional blue
        subtotal_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')      # Light blue

        for section in bank_sections:
            row += 2  # Add spacing

            bank_code = section['code']
            doc_type = section['type']
            section_title = section['title']

            # Get documents for this section
            documents = sections_data.get((bank_code, doc_type), [])

            # Get first document with bank account info (if exists)
            sample_account = None
            if documents:
                for doc in documents:
                    if doc.company_bank_account:
                        sample_account = doc.company_bank_account
                        break

            # ===========================================================================
            # SECTION HEADER ROW (Yellow/Orange background)
            # ===========================================================================
            ws.merge_cells(f'A{row}:G{row}')
            header_cell = ws[f'A{row}']

            if sample_account:
                header_cell.value = f"Transfer by Account: {sample_account.company_name}, Account Number {sample_account.account_number} ({sample_account.currency}) {sample_account.bank}"
            else:
                header_cell.value = f"Transfer by Account: {section['name']} (No account assigned)"

            header_cell.font = Font(name='Calibri', size=11, bold=True, color='000000')
            header_cell.alignment = left_alignment
            header_cell.fill = section_header_fill
            header_cell.border = border

            # Section title on right (H column)
            title_cell = ws[f'H{row}']
            title_cell.value = section_title
            title_cell.font = Font(name='Calibri', size=12, bold=True, color='000000')
            title_cell.alignment = center_alignment
            title_cell.fill = section_header_fill
            title_cell.border = border

            row += 1

            # ===========================================================================
            # TABLE HEADERS
            # ===========================================================================
            table_headers = [
                ('A', 'No'),
                ('B', 'PV No.' if doc_type == 'PV' else 'PF No.'),
                ('C', 'Supplier'),
                ('D', 'Description'),
                ('E', 'Amount'),
                ('F', 'Receiver Bank Account\nAccount Name'),
                ('G', 'Receiver Bank Account\nAccount Number'),
                ('H', 'Remark')
            ]

            for col, header_text in table_headers:
                cell = ws[f'{col}{row}']
                cell.value = header_text
                cell.font = Font(name='Calibri', size=10, bold=True, color='000000')
                cell.alignment = center_alignment
                cell.fill = table_header_fill
                cell.border = border

            row += 1

            # ===========================================================================
            # DATA ROWS
            # ===========================================================================
            section_totals = {'USD': Decimal('0'), 'KHR': Decimal('0'), 'THB': Decimal('0')}

            if documents:
                for idx, doc in enumerate(documents, 1):
                    # Determine if this is a voucher or form
                    is_voucher = hasattr(doc, 'pv_number')
                    doc_number = doc.pv_number if is_voucher else doc.pf_number

                    # Calculate totals for all currencies
                    totals = doc.calculate_grand_total()

                    # Build amount display string with all currencies
                    amount_parts = []
                    for currency in ['USD', 'KHR', 'THB']:
                        if currency in totals and totals[currency] > 0:
                            amount_parts.append(f"{currency} {totals[currency]:,.2f}")
                            section_totals[currency] += totals[currency]
                            grand_totals[currency] += totals[currency]

                    amount_display = '\n'.join(amount_parts) if amount_parts else '0.00'

                    # Combine all line item descriptions
                    descriptions = [item.description for item in doc.line_items.all()]
                    combined_description = '\n'.join(descriptions) if descriptions else ''

                    # Receiver bank info
                    receiver_account_name = doc.bank_name or '-'
                    receiver_account_number = doc.bank_account_number or '-'

                    # Column A: No
                    ws[f'A{row}'] = idx
                    ws[f'A{row}'].font = data_font
                    ws[f'A{row}'].alignment = center_alignment
                    ws[f'A{row}'].border = border

                    # Column B: Doc Number
                    ws[f'B{row}'] = doc_number or 'DRAFT'
                    ws[f'B{row}'].font = data_font
                    ws[f'B{row}'].alignment = left_alignment
                    ws[f'B{row}'].border = border

                    # Column C: Supplier
                    ws[f'C{row}'] = doc.payee_name
                    ws[f'C{row}'].font = data_font
                    ws[f'C{row}'].alignment = left_alignment
                    ws[f'C{row}'].border = border

                    # Column D: Description
                    ws[f'D{row}'] = combined_description
                    ws[f'D{row}'].font = data_font
                    ws[f'D{row}'].alignment = left_alignment
                    ws[f'D{row}'].border = border

                    # Column E: Amount (with all currencies)
                    ws[f'E{row}'] = amount_display
                    ws[f'E{row}'].font = data_font
                    ws[f'E{row}'].alignment = right_alignment
                    ws[f'E{row}'].border = border

                    # Column F: Receiver Account Name
                    ws[f'F{row}'] = receiver_account_name
                    ws[f'F{row}'].font = data_font
                    ws[f'F{row}'].alignment = left_alignment
                    ws[f'F{row}'].border = border

                    # Column G: Receiver Account Number
                    ws[f'G{row}'] = receiver_account_number
                    ws[f'G{row}'].font = data_font
                    ws[f'G{row}'].alignment = left_alignment
                    ws[f'G{row}'].border = border

                    # Column H: Remark (blank)
                    ws[f'H{row}'] = ''
                    ws[f'H{row}'].font = data_font
                    ws[f'H{row}'].alignment = left_alignment
                    ws[f'H{row}'].border = border

                    row += 1
            else:
                # No documents - add 3 empty rows
                for _ in range(3):
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                        ws[f'{col}{row}'] = ''
                        ws[f'{col}{row}'].border = border
                    row += 1

            # ===========================================================================
            # SUBTOTAL ROW
            # ===========================================================================
            ws.merge_cells(f'A{row}:D{row}')
            subtotal_label = ws[f'A{row}']
            subtotal_label.value = f"Subtotal - {section_title}"
            subtotal_label.font = Font(name='Calibri', size=11, bold=True)
            subtotal_label.alignment = right_alignment
            subtotal_label.fill = subtotal_fill
            subtotal_label.border = border

            # Build subtotal display with all currencies
            subtotal_parts = []
            for currency in ['USD', 'KHR', 'THB']:
                if section_totals[currency] > 0:
                    subtotal_parts.append(f"{currency} {section_totals[currency]:,.2f}")
            subtotal_display = '\n'.join(subtotal_parts) if subtotal_parts else '0.00'

            ws[f'E{row}'] = subtotal_display
            ws[f'E{row}'].font = Font(name='Calibri', size=11, bold=True)
            ws[f'E{row}'].alignment = right_alignment
            ws[f'E{row}'].fill = subtotal_fill
            ws[f'E{row}'].border = border

            # Empty cells for subtotal row
            for col in ['F', 'G', 'H']:
                ws[f'{col}{row}'] = ''
                ws[f'{col}{row}'].fill = subtotal_fill
                ws[f'{col}{row}'].border = border

            row += 1

        # ===========================================================================
        # GRAND TOTAL ROW
        # ===========================================================================
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        grand_total_label = ws[f'A{row}']
        grand_total_label.value = "GRAND TOTAL"
        grand_total_label.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        grand_total_label.alignment = right_alignment
        grand_total_label.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        grand_total_label.border = border

        # Build grand total display with all currencies
        grand_total_parts = []
        for currency in ['USD', 'KHR', 'THB']:
            if grand_totals[currency] > 0:
                grand_total_parts.append(f"{currency} {grand_totals[currency]:,.2f}")
        grand_total_display = '\n'.join(grand_total_parts) if grand_total_parts else '0.00'

        ws[f'E{row}'] = grand_total_display
        ws[f'E{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        ws[f'E{row}'].alignment = right_alignment
        ws[f'E{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws[f'E{row}'].border = border

        for col in ['F', 'G', 'H']:
            ws[f'{col}{row}'] = ''
            ws[f'{col}{row}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws[f'{col}{row}'].border = border

        row += 1

        # ===========================================================================
        # SIGNATURE FOOTER (3 COLUMNS)
        # ===========================================================================
        row += 3  # Add spacing

        # Get current user for "Prepared By"
        from django.contrib.auth import get_user_model
        User = get_user_model()

        # Prepared By (Columns A-B)
        ws.merge_cells(f'A{row}:B{row}')
        prepared_cell = ws[f'A{row}']
        prepared_cell.value = "Prepared By:"
        prepared_cell.font = Font(name='Calibri', size=11, bold=True)
        prepared_cell.alignment = left_alignment

        ws.merge_cells(f'A{row+1}:B{row+1}')
        prepared_name = ws[f'A{row+1}']
        prepared_name.value = "___________________________"
        prepared_name.font = data_font
        prepared_name.alignment = center_alignment

        ws.merge_cells(f'A{row+2}:B{row+2}')
        prepared_date = ws[f'A{row+2}']
        prepared_date.value = "Date: ............"
        prepared_date.font = data_font
        prepared_date.alignment = center_alignment

        # Checked By (Columns C-E)
        ws.merge_cells(f'C{row}:E{row}')
        checked_cell = ws[f'C{row}']
        checked_cell.value = "Checked By: Finance Manager"
        checked_cell.font = Font(name='Calibri', size=11, bold=True)
        checked_cell.alignment = center_alignment

        ws.merge_cells(f'C{row+1}:E{row+1}')
        checked_name = ws[f'C{row+1}']
        checked_name.value = "___________________________"
        checked_name.font = data_font
        checked_name.alignment = center_alignment

        ws.merge_cells(f'C{row+2}:E{row+2}')
        checked_date = ws[f'C{row+2}']
        checked_date.value = "Date: ............"
        checked_date.font = data_font
        checked_date.alignment = center_alignment

        # Approved By (Columns F-H)
        ws.merge_cells(f'F{row}:H{row}')
        approved_cell = ws[f'F{row}']
        approved_cell.value = "Approved By: Managing Director"
        approved_cell.font = Font(name='Calibri', size=11, bold=True)
        approved_cell.alignment = center_alignment

        ws.merge_cells(f'F{row+1}:H{row+1}')
        approved_name = ws[f'F{row+1}']
        approved_name.value = "___________________________"
        approved_name.font = data_font
        approved_name.alignment = center_alignment

        ws.merge_cells(f'F{row+2}:H{row+2}')
        approved_date = ws[f'F{row+2}']
        approved_date.value = "Date: ............"
        approved_date.font = data_font
        approved_date.alignment = center_alignment

        # ===========================================================================
        # PAGE SETUP
        # ===========================================================================
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = False

        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        ws.print_area = f'A1:H{row+2}'
        ws.print_options.gridLines = False

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def export_to_pdf(self):
        """Export filtered data to PDF with professional formatting"""
        if self.vouchers is None or self.forms is None:
            self.apply_filters()

        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)

        # Container for elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=12,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1F4E78'),
            spaceAfter=10,
            spaceBefore=10
        )

        # Title
        title = Paragraph("Payment Vouchers & Forms Report", title_style)
        elements.append(title)

        # Report info
        info_text = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        info = Paragraph(info_text, styles['Normal'])
        elements.append(info)
        elements.append(Spacer(1, 12))

        # Filter info
        if self.filters.get('date_from') or self.filters.get('date_to') or self.filters.get('status'):
            filter_parts = []
            if self.filters.get('date_from'):
                filter_parts.append(f"From: {self.filters['date_from']}")
            if self.filters.get('date_to'):
                filter_parts.append(f"To: {self.filters['date_to']}")
            if self.filters.get('status') and self.filters['status'] != 'ALL':
                filter_parts.append(f"Status: {self.filters['status']}")

            filter_text = "Filters: " + " | ".join(filter_parts)
            filters_para = Paragraph(filter_text, styles['Normal'])
            elements.append(filters_para)
            elements.append(Spacer(1, 12))

        # Table data
        data = [['No', 'Type', 'Supplier', 'Description', 'Amount', 'Transfer by Account', 'Remark']]

        # Add vouchers
        for voucher in self.vouchers:
            totals = voucher.calculate_grand_total()
            descriptions = " | ".join([item.description for item in voucher.line_items.all()[:2]])[:60]

            # Format amount with currency
            amount_parts = []
            for currency in ['USD', 'KHR', 'THB']:
                if totals.get(currency, 0) > 0:
                    symbols = {'USD': '$', 'KHR': '៛', 'THB': '฿'}
                    amount_parts.append(f"{symbols[currency]}{totals[currency]:,.2f}")
            amount_display = " + ".join(amount_parts) if amount_parts else "$0.00"

            # Get transfer account info - prioritize company_bank_account
            if voucher.company_bank_account:
                transfer_account = voucher.company_bank_account.get_display_name()[:50]
            elif voucher.bank_name or voucher.bank_account_number:
                # Fallback to manual entry
                parts = []
                if voucher.bank_name:
                    parts.append(voucher.bank_name)
                if voucher.bank_account_number:
                    parts.append(f"Acc '{voucher.bank_account_number}")
                if voucher.bank_address:
                    parts.append(voucher.bank_address)
                transfer_account = ", ".join(parts)[:50]
            else:
                transfer_account = ''

            data.append([
                voucher.pv_number or 'DRAFT',
                'PV',
                voucher.payee_name[:25],
                descriptions,
                amount_display[:20],
                transfer_account,
                ''  # Blank remark
            ])

        # Add forms
        for form in self.forms:
            totals = form.calculate_grand_total()
            descriptions = " | ".join([item.description for item in form.line_items.all()[:2]])[:60]

            # Format amount with currency
            amount_parts = []
            for currency in ['USD', 'KHR', 'THB']:
                if totals.get(currency, 0) > 0:
                    symbols = {'USD': '$', 'KHR': '៛', 'THB': '฿'}
                    amount_parts.append(f"{symbols[currency]}{totals[currency]:,.2f}")
            amount_display = " + ".join(amount_parts) if amount_parts else "$0.00"

            # Get transfer account info - prioritize company_bank_account
            if form.company_bank_account:
                transfer_account = form.company_bank_account.get_display_name()[:50]
            elif form.bank_name or form.bank_account_number:
                # Fallback to manual entry
                parts = []
                if form.bank_name:
                    parts.append(form.bank_name)
                if form.bank_account_number:
                    parts.append(f"Acc '{form.bank_account_number}")
                if form.bank_address:
                    parts.append(form.bank_address)
                transfer_account = ", ".join(parts)[:50]
            else:
                transfer_account = ''

            data.append([
                form.pf_number or 'DRAFT',
                'PF',
                form.payee_name[:25],
                descriptions,
                amount_display[:20],
                transfer_account,
                ''  # Blank remark
            ])

        # Create table with adjusted column widths for landscape A4
        table = Table(data, colWidths=[0.9*inch, 0.4*inch, 1.3*inch, 2.3*inch, 1.1*inch, 2.5*inch, 1.5*inch])

        # Table style
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 24))

        # Summary section
        elements.append(Paragraph("Summary Statistics", heading_style))
        elements.append(Spacer(1, 12))

        stats = self.get_summary_stats()

        summary_data = [
            ['Total Documents:', str(stats['total_documents'])],
            ['Payment Vouchers (PV):', str(stats['total_vouchers'])],
            ['Payment Forms (PF):', str(stats['total_forms'])],
            ['', ''],
            ['Total Amount (USD):', f"${stats['by_currency']['USD']:,.2f}"],
            ['Total Amount (KHR):', f"៛{stats['by_currency']['KHR']:,.2f}"],
            ['Total Amount (THB):', f"฿{stats['by_currency']['THB']:,.2f}"],
        ]

        summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, 2), 1, colors.grey),
            ('GRID', (0, 4), (-1, -1), 1, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4F8')),
        ]))

        elements.append(summary_table)

        # Build PDF
        doc.build(elements)

        buffer.seek(0)
        return buffer
