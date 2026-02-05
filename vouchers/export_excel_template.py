"""
Excel Export View with Custom Template Format
Exports Payment Vouchers to Excel matching the exact template specification
"""

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import PaymentVoucher, PaymentForm


@login_required
def export_excel_template_view(request):
    """
    Export Payment Vouchers to Excel with custom template format.

    GET Parameters:
        - date_from: Filter start date (YYYY-MM-DD)
        - date_to: Filter end date (YYYY-MM-DD)
        - status: Filter by status (default: APPROVED)
        - doc_type: PV or PF or ALL
        - creator: Filter by creator user ID
        - department: Filter by department name
        - payee_name: Filter by payee name (partial match)
        - bank_account: Filter by specific company bank account ID
    """

    # ===========================================================================
    # 1. APPLY FILTERS
    # ===========================================================================

    vouchers_qs = PaymentVoucher.objects.all()
    forms_qs = PaymentForm.objects.all()

    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if date_from:
        vouchers_qs = vouchers_qs.filter(payment_date__gte=date_from)
        forms_qs = forms_qs.filter(payment_date__gte=date_from)

    if date_to:
        vouchers_qs = vouchers_qs.filter(payment_date__lte=date_to)
        forms_qs = forms_qs.filter(payment_date__lte=date_to)

    # Status filter - default to APPROVED only
    status = request.GET.get('status', 'APPROVED')
    if status and status != 'ALL':
        vouchers_qs = vouchers_qs.filter(status=status)
        forms_qs = forms_qs.filter(status=status)

    # Creator filter
    creator_id = request.GET.get('creator')
    if creator_id:
        vouchers_qs = vouchers_qs.filter(created_by_id=creator_id)
        forms_qs = forms_qs.filter(created_by_id=creator_id)

    # Department filter
    department_name = request.GET.get('department')
    if department_name:
        vouchers_qs = vouchers_qs.filter(line_items__department__name=department_name).distinct()
        forms_qs = forms_qs.filter(line_items__department__name=department_name).distinct()

    # Payee filter
    payee_name = request.GET.get('payee_name')
    if payee_name:
        vouchers_qs = vouchers_qs.filter(payee_name__icontains=payee_name)
        forms_qs = forms_qs.filter(payee_name__icontains=payee_name)

    # Bank account filter
    bank_account_id = request.GET.get('bank_account')
    if bank_account_id:
        vouchers_qs = vouchers_qs.filter(company_bank_account_id=bank_account_id)
        forms_qs = forms_qs.filter(company_bank_account_id=bank_account_id)

    # Document type filter
    doc_type = request.GET.get('doc_type', 'ALL')
    if doc_type == 'PV':
        forms_qs = PaymentForm.objects.none()
    elif doc_type == 'PF':
        vouchers_qs = PaymentVoucher.objects.none()

    # Prefetch related data
    vouchers = vouchers_qs.select_related('created_by', 'company_bank_account').prefetch_related('line_items')
    forms = forms_qs.select_related('created_by', 'company_bank_account').prefetch_related('line_items')

    # Combine vouchers and forms into a single list
    all_documents = list(vouchers) + list(forms)

    # Sort by payment date
    all_documents.sort(key=lambda x: x.payment_date)

    # ===========================================================================
    # 2. CREATE WORKBOOK
    # ===========================================================================

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Voucher Summary"

    # ===========================================================================
    # 3. SET COLUMN WIDTHS (EXACT MATCH TO TEMPLATE)
    # ===========================================================================

    column_widths = {
        'A': 3.71,   # No
        'B': 15.14,  # PV No
        'C': 35.86,  # Supplier
        'D': 71.57,  # Description
        'E': 23.71,  # Amount
        'F': 32.86,  # Account Name
        'G': 25.86,  # Account Number
        'H': 22.0    # Remark
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ===========================================================================
    # 4. DEFINE STYLES
    # ===========================================================================

    # Font styles
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')
    company_font = Font(name='Calibri', size=14, bold=True, color='000000')
    title_font = Font(name='Calibri', size=13, bold=True, color='000000')
    data_font = Font(name='Calibri', size=11, color='000000')

    # Fill styles
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Alignment styles
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    right_alignment = Alignment(horizontal='right', vertical='center')

    # ===========================================================================
    # 5. ROW 1: COMPANY NAME
    # ===========================================================================

    ws.merge_cells('A1:H1')
    company_cell = ws['A1']
    # TODO: Customize company name here
    company_cell.value = "Phat Phnom Penh  Co.,LTD  (Garden City Water Park)"
    company_cell.font = company_font
    company_cell.alignment = center_alignment

    # ===========================================================================
    # 6. ROW 2: TITLE
    # ===========================================================================

    ws.merge_cells('A2:H2')
    title_cell = ws['A2']
    title_cell.value = "Summary Payment Voucher"
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    # ===========================================================================
    # 7. ROW 3: ACCOUNT NUMBER AND DATE
    # ===========================================================================

    # Account number (Column D)
    account_cell = ws['D3']
    # TODO: This should be dynamic based on filtered bank account
    # For now, we'll use the first document's account if available
    if all_documents and hasattr(all_documents[0], 'company_bank_account') and all_documents[0].company_bank_account:
        account_cell.value = all_documents[0].company_bank_account.account_number
    else:
        account_cell.value = "N/A"  # Fallback if no account selected
    account_cell.font = data_font
    account_cell.alignment = left_alignment

    # "Date" label (Column G)
    date_label_cell = ws['G3']
    date_label_cell.value = "Date"
    date_label_cell.font = header_font
    date_label_cell.alignment = right_alignment

    # Current date (Column H)
    date_cell = ws['H3']
    date_cell.value = timezone.now().strftime('%Y-%m-%d')
    date_cell.font = data_font
    date_cell.alignment = left_alignment

    # ===========================================================================
    # 8. ROW 4: TRANSFER BY ACCOUNT
    # ===========================================================================

    ws.merge_cells('A4:H4')
    transfer_cell = ws['A4']

    # Get transfer account info from first document
    if all_documents and hasattr(all_documents[0], 'company_bank_account') and all_documents[0].company_bank_account:
        bank_acc = all_documents[0].company_bank_account
        transfer_cell.value = f"Transfer by Account:  {bank_acc.company_name}, Account Number '{bank_acc.account_number} ({bank_acc.currency}) {bank_acc.bank}"
    else:
        transfer_cell.value = "Transfer by Account:  [Not Specified]"

    transfer_cell.font = Font(name='Calibri', size=11, bold=False)
    transfer_cell.alignment = left_alignment

    # ===========================================================================
    # 9. ROWS 5-6: COLUMN HEADERS (WITH MERGED CELLS)
    # ===========================================================================

    # Define header structure
    headers_row5 = [
        ('A5', 'No'),
        ('B5', 'PV No.'),
        ('C5', 'Supplier'),
        ('D5', 'Description'),
        ('E5', 'Amount'),
        ('F5', 'Receiver Bank Account'),  # This will be merged F5:G5
        ('H5', 'Remark')
    ]

    headers_row6 = [
        ('F6', 'Account Name'),
        ('G6', 'Account Number')
    ]

    # Merge cells for row 5
    merge_ranges = [
        'A5:A6',  # No
        'B5:B6',  # PV No
        'C5:C6',  # Supplier
        'D5:D6',  # Description
        'E5:E6',  # Amount
        'F5:G5',  # Receiver Bank Account (merged horizontally)
        'H5:H6'   # Remark
    ]

    for merge_range in merge_ranges:
        ws.merge_cells(merge_range)

    # Apply row 5 headers
    for cell_ref, value in headers_row5:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Apply row 6 sub-headers
    for cell_ref, value in headers_row6:
        cell = ws[cell_ref]
        cell.value = value
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Apply borders to all merged header cells
    for row in [5, 6]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f'{col}{row}']
            if not cell.value:  # Only apply if not already set
                cell.border = thin_border
                cell.fill = header_fill

    # ===========================================================================
    # 10. DATA ROWS (STARTING FROM ROW 7)
    # ===========================================================================

    current_row = 7
    total_amount = 0

    for idx, doc in enumerate(all_documents, start=1):
        # Get document type
        doc_type_str = 'PV' if isinstance(doc, PaymentVoucher) else 'PF'

        # Get voucher/form number
        doc_number = doc.pv_number if isinstance(doc, PaymentVoucher) else doc.pf_number

        # Combine all line item descriptions
        descriptions = []
        for item in doc.line_items.all():
            descriptions.append(item.description)
        combined_description = '\n'.join(descriptions) if descriptions else ''

        # Calculate total amount
        totals = doc.calculate_grand_total()
        # For simplicity, we'll show USD amount (you can customize this)
        amount = float(totals.get('USD', 0))
        total_amount += amount

        # Get bank account information
        if doc.company_bank_account:
            account_name = f"{doc.company_bank_account.company_name}"
            account_number = doc.company_bank_account.account_number
        elif doc.bank_name or doc.bank_account_number:
            # Fallback to manual entry
            account_name = doc.bank_name or ''
            account_number = doc.bank_account_number or ''
        else:
            account_name = ''
            account_number = ''

        # Column A: Sequential number
        cell_a = ws[f'A{current_row}']
        cell_a.value = idx
        cell_a.font = data_font
        cell_a.alignment = center_alignment
        cell_a.border = thin_border

        # Column B: Voucher number
        cell_b = ws[f'B{current_row}']
        cell_b.value = doc_number or 'DRAFT'
        cell_b.font = data_font
        cell_b.alignment = left_alignment
        cell_b.border = thin_border

        # Column C: Supplier/Payee
        cell_c = ws[f'C{current_row}']
        cell_c.value = doc.payee_name
        cell_c.font = data_font
        cell_c.alignment = left_alignment
        cell_c.border = thin_border

        # Column D: Description
        cell_d = ws[f'D{current_row}']
        cell_d.value = combined_description
        cell_d.font = data_font
        cell_d.alignment = left_alignment
        cell_d.border = thin_border

        # Column E: Amount (right-aligned, formatted)
        cell_e = ws[f'E{current_row}']
        cell_e.value = amount
        cell_e.font = data_font
        cell_e.alignment = right_alignment
        cell_e.border = thin_border
        cell_e.number_format = '#,##0.00'

        # Column F: Account Name
        cell_f = ws[f'F{current_row}']
        cell_f.value = account_name
        cell_f.font = data_font
        cell_f.alignment = left_alignment
        cell_f.border = thin_border

        # Column G: Account Number
        cell_g = ws[f'G{current_row}']
        cell_g.value = account_number
        cell_g.font = data_font
        cell_g.alignment = left_alignment
        cell_g.border = thin_border

        # Column H: Remark (blank for manual entry)
        cell_h = ws[f'H{current_row}']
        cell_h.value = ''
        cell_h.font = data_font
        cell_h.alignment = left_alignment
        cell_h.border = thin_border

        current_row += 1

    # ===========================================================================
    # 11. TOTAL ROW
    # ===========================================================================

    # Column D: "Total:" label
    total_label_cell = ws[f'D{current_row}']
    total_label_cell.value = "Total:"
    total_label_cell.font = header_font
    total_label_cell.alignment = right_alignment
    total_label_cell.border = thin_border

    # Column E: Total amount
    total_amount_cell = ws[f'E{current_row}']
    total_amount_cell.value = total_amount
    total_amount_cell.font = header_font
    total_amount_cell.alignment = right_alignment
    total_amount_cell.border = thin_border
    total_amount_cell.number_format = '#,##0.00'

    # Apply borders to other cells in total row
    for col in ['A', 'B', 'C', 'F', 'G', 'H']:
        cell = ws[f'{col}{current_row}']
        cell.border = thin_border

    # ===========================================================================
    # 12. SIGNATURE SECTION (3 ROWS AFTER TOTAL)
    # ===========================================================================

    signature_row = current_row + 3

    # Merge B:C for "Prepared by"
    ws.merge_cells(f'B{signature_row}:C{signature_row}')
    prepared_cell = ws[f'B{signature_row}']
    prepared_cell.value = "Prepared by"
    prepared_cell.font = header_font
    prepared_cell.alignment = center_alignment

    # You can add more signature sections here if needed
    # For example: "Checked by", "Approved by", etc.

    # ===========================================================================
    # 13. GENERATE RESPONSE
    # ===========================================================================

    # Create filename with timestamp
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Payment_Voucher_Summary_{timestamp}.xlsx'

    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
